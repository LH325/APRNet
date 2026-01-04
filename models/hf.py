import torch
import torch.nn as nn
from layers.RevIN import RevIN
import numpy as np
from einops import rearrange
import pandas as pd
import torch.fft
import torch.nn.functional as F
import einops
from layers.vlm_manager import VLMManager
from einops import rearrange
import math
import numpy as np
import matplotlib.pyplot as plt
import os

# ========================= 跨结构因果注意力：严格+软因果 =========================
class CausalMultiheadAttentionWrapper(nn.Module):
    def __init__(self, embed_dim=1024, num_heads=8, dropout=0.4, soft_penalty=-4.0):
        super().__init__()
        self.embed_dim = embed_dim
        self.num_heads = num_heads
        self.soft_penalty = soft_penalty
        self.mha = nn.MultiheadAttention(embed_dim=embed_dim, num_heads=num_heads, dropout=dropout, batch_first=True)

    def _mask(self, Tq, Tk, device, soft=False):
        i = torch.arange(Tq, device=device).unsqueeze(1)
        j = torch.arange(Tk, device=device).unsqueeze(0)
        future = (j > i)
        m = torch.zeros(Tq, Tk, device=device)
        if soft:
            m = m.masked_fill(future, float(self.soft_penalty))
        else:
            m = m.masked_fill(future, float('-inf'))
        return m

    def forward(self, x, y):
        B, Tq, D = x.shape
        Tk = y.size(1)
        m_strict = self._mask(Tq, Tk, x.device, soft=False)
        m_soft = self._mask(Tq, Tk, x.device, soft=True)

        out_strict, attn_strict = self.mha(query=x, key=y, value=y, attn_mask=m_strict, need_weights=True, average_attn_weights=False)

        _, attn_soft = self.mha(query=x, key=y, value=y, attn_mask=m_soft, need_weights=True, average_attn_weights=False)
        return out_strict, attn_strict, attn_soft


# ========================= 位置编码模块 =========================
class SinusoidalPositionEncoding(nn.Module):
    def __init__(self, channels=21, height=32, width=32, temperature=100000.0):
        super().__init__()
        self.channels = channels
        self.height = height
        self.width = width
        self.temperature = temperature
        self.register_buffer('pos_encoding', self._create_pos_encoding())

    def _create_pos_encoding(self):
        y_pos = torch.arange(self.height).float()
        x_pos = torch.arange(self.width).float()
        y_pos = y_pos / (self.height - 1) if self.height > 1 else y_pos
        x_pos = x_pos / (self.width - 1) if self.width > 1 else x_pos
        grid_y, grid_x = torch.meshgrid(y_pos, x_pos, indexing='ij')
        half_dim = self.channels // 2
        dim_t = torch.arange(half_dim).float()
        dim_t = self.temperature ** (2 * dim_t / max(half_dim, 1))
        pos_x = grid_x.unsqueeze(-1) / dim_t
        pos_y = grid_y.unsqueeze(-1) / dim_t
        pos_x = torch.stack([pos_x.sin(), pos_x.cos()], dim=-1).flatten(-2)
        pos_y = torch.stack([pos_y.sin(), pos_y.cos()], dim=-1).flatten(-2)
        pos_encoding = pos_x + pos_y
        if pos_encoding.shape[-1] > self.channels:
            pos_encoding = pos_encoding[..., :self.channels]
        elif pos_encoding.shape[-1] < self.channels:
            padding = torch.zeros(*pos_encoding.shape[:-1], self.channels - pos_encoding.shape[-1])
            pos_encoding = torch.cat([pos_encoding, padding], dim=-1)
        pos_encoding = pos_encoding.permute(2, 0, 1).unsqueeze(0)
        return pos_encoding

    def forward(self, x):
        return x + self.pos_encoding


class PositionalEncoding(nn.Module):
    def __init__(self, d_model, max_len=5000):
        super().__init__()
        pe = torch.zeros(max_len, d_model)
        position = torch.arange(0, max_len, dtype=torch.float).unsqueeze(1)
        div_term = torch.exp(torch.arange(0, d_model, 2).float() * (-np.log(100000.0) / d_model))
        pe[:, 0::2] = torch.sin(position * div_term)
        pe[:, 1::2] = torch.cos(position * div_term)
        self.register_buffer('pe', pe.unsqueeze(0))

    def forward(self, x):
        return x + self.pe[:, :x.size(1), :]


# ========================= 1D -> 2D 简化映射 =========================
def time_series_to_simple_image(x_enc, image_size, context_len, periodicity):
    B, seq_len, nvars = x_enc.shape
    pad_left = 0
    if context_len % periodicity != 0:
        pad_left = periodicity - context_len % periodicity
    x_enc = einops.rearrange(x_enc, 'b s n -> b n s')
    x_pad = F.pad(x_enc, (pad_left, 0), mode='replicate')
    x_2d = einops.rearrange(x_pad, 'b n (p f) -> (b n) 1 f p', f=periodicity)
    x_resized_2d = F.interpolate(x_2d, size=(image_size, image_size), mode='bilinear', align_corners=False)
    images = einops.repeat(x_resized_2d, 'b 1 h w -> b c h w', c=3)
    video = einops.rearrange(images, '(b n) c h w -> b n c h w', b=B, n=nvars)
    images = video.mean(dim=1)
    return images, images


# ========================= 融合模块 =========================
class temporal(nn.Module):
    def __init__(self, seq_len, pred_len, enc_in, embed, hidden_size=1024):
        super().__init__()
        self.temporal_ = nn.Sequential(
            nn.Linear(embed, hidden_size), nn.ReLU(), nn.Linear(hidden_size, embed)
        )

    def forward(self, x):
        return x + self.temporal_(x)


class imagesorocess(nn.Module):
    def __init__(self, seq_len, pred_len, enc_in, embed, hidden_size=1024):
        super().__init__()
        self.images = nn.Sequential(
            nn.Linear(embed, hidden_size), nn.ReLU(), nn.Linear(hidden_size, embed)
        )

    def forward(self, x):
        return x + self.images(x)


class FusionModule(nn.Module):
    def __init__(self, seq_len, pred_len, enc_in, embed):
        super().__init__()
        self.model = temporal(seq_len, pred_len, enc_in, embed)
        self.modeli = imagesorocess(seq_len, pred_len, enc_in, embed)
        self.gate = nn.Sequential(
            nn.Linear(embed * 2, embed), nn.GELU(), nn.Linear(embed, 2), nn.Softmax(dim=-1)
        )

    def forward(self, x, img):
        temporal_out = self.model(x)
        multimodal_features = self.modeli(img)
        fuse = torch.cat([temporal_out, multimodal_features], dim=-1)
        gate_weights_seq = self.gate(fuse)
        fused = gate_weights_seq[:, :, 0:1] * temporal_out + gate_weights_seq[:, :, 1:2] * multimodal_features
        return fused


# ========================= 主模型：CDIA =========================
class Model(nn.Module):
    def __init__(self, configs):
        super(Model, self).__init__()
        configs.log_excel_path = "loss_log.xlsx"
        self.vlm_manager = VLMManager(configs)
        self.seq_len = configs.seq_len
        self.pred_len = configs.pred_len
        self.revin_layer = RevIN(configs.enc_in, affine=False, subtract_last=False)

        self.embed_size = configs.embed_size
        self.hidden_size = configs.hidden_size
        self.enc_in = configs.enc_in
        self.criterion = nn.MSELoss()
        self.fc1 = nn.Sequential(
            nn.Linear(self.embed_size, self.hidden_size),
            nn.ReLU(),
            nn.Linear(self.hidden_size, self.hidden_size),
            nn.ReLU(),
            nn.Dropout(0.4),
            nn.Linear(self.hidden_size, self.pred_len)
        )
        self.attn1 = FusionModule(self.seq_len, self.pred_len, self.enc_in, self.embed_size)
        self.V_embedding = nn.Sequential(nn.Conv2d(self.enc_in, self.enc_in, 3, 3), nn.ReLU())
        self.I_embedding = nn.Sequential(nn.Conv2d(self.enc_in, self.enc_in, 3, 3), nn.ReLU())
        self.layernormV = nn.LayerNorm(self.embed_size)
        self.layernormI = nn.LayerNorm(self.embed_size)
        self.layernormT = nn.LayerNorm(self.embed_size)
        self.layernormT1 = nn.LayerNorm(self.embed_size)
        self.T_embedding = nn.Linear(self.seq_len, self.embed_size)
        self.pos_emb = PositionalEncoding(d_model=self.embed_size)
        self.img_pos_emb = SinusoidalPositionEncoding()

        soft_penalty = getattr(configs, 'soft_penalty', -4.0)
        self.cross = CausalMultiheadAttentionWrapper(self.embed_size, 8, 0.4, soft_penalty)
        self.crossv = CausalMultiheadAttentionWrapper(self.embed_size, 8, 0.4, soft_penalty)
        self.mlp = nn.Sequential(nn.Linear(self.embed_size * 2, self.embed_size), nn.ReLU(), nn.Linear(self.embed_size, self.embed_size))
        self.mi_tau = getattr(configs, 'mi_tau', 0.07)
        self.lambda_causal = getattr(configs, 'lambda_causal', 1.0)
        self.alpha_align = getattr(configs, 'alpha_align', 0.5)
        self.eta_future = getattr(configs, 'eta_future', 1.0)
        self.w_cdia = getattr(configs, 'w_cdia', 0.1)
        self.align_proj = nn.Linear(self.embed_size, self.embed_size)
        self.last_losses = {}

        # ===== Excel日志配置 =====
        self.log_excel_path = getattr(configs, 'log_excel_path', None)
        self._excel_initialized = False
        self._log_step = 0

    # ---------- Excel初始化 ----------
    def _init_excel_if_needed(self, extra_keys=None):
        if self.log_excel_path is None or self._excel_initialized:
            return
        cols = ['step', 'time'] + list(self.last_losses.keys())
        if extra_keys:
            for k in extra_keys:
                if k not in cols:
                    cols.append(k)
        pd.DataFrame(columns=cols).to_excel(self.log_excel_path, index=False)
        self._excel_initialized = True

    # ---------- Excel追加 ----------
    def _append_losses_to_excel(self, step, extra=None):
        if self.log_excel_path is None:
            return
        from datetime import datetime
        try:
            self._init_excel_if_needed(extra_keys=(extra.keys() if extra else None))
            from openpyxl import load_workbook
            wb = load_workbook(self.log_excel_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            row_dict = {'step': step, 'time': datetime.now().isoformat(timespec='seconds')}
            row_dict.update(self.last_losses)
            if extra:
                row_dict.update(extra)
            ws.append([row_dict.get(h, None) for h in headers])
            wb.save(self.log_excel_path)
        except Exception as e:
            print(f"[ExcelLog] append failed: {e}")

    # ---------- Normalize ----------
    def _normalize_images(self, images):
        min_vals = images.reshape(images.size(0), -1).min(dim=1, keepdim=True)[0].view(-1, 1, 1, 1)
        max_vals = images.reshape(images.size(0), -1).max(dim=1, keepdim=True)[0].view(-1, 1, 1, 1)
        images = (images - min_vals) / (max_vals - min_vals + 1e-5)
        images = (images * 255).clamp(0, 255).to(torch.uint8)
        return images

    def vision_augmented_learner(self, x_enc, image_size, context_len, periodicity):
        images, videos = time_series_to_simple_image(x_enc, image_size, context_len, periodicity)
        return self._normalize_images(images), self._normalize_images(videos)

    # ---------- 前向 ----------
    def forward(self, x, x_mark_enc, x_dec, x_mark_dec, y=None, mask=None):
        z = self.revin_layer(x, 'norm')
        x = z
        device = x.device
        images, videos = self.vision_augmented_learner(z, 96, self.seq_len, 96)
        videos = videos.float().to(device)
        videos = videos.sum(1).unsqueeze(1).expand(-1, self.enc_in, -1, -1)
        videos_embeddings = self.layernormV(rearrange(self.img_pos_emb(self.V_embedding(videos)), 'B C W H->B C (W H)'))
        images = images.float().to(device)
        images = images.sum(1).unsqueeze(1).expand(-1, self.enc_in, -1, -1)
        vision_embeddings = self.layernormI(rearrange(self.img_pos_emb(self.I_embedding(images)), 'B C W H->B C (W H)'))
        F_repr, A_gl_strict, A_gl_soft = self.cross(vision_embeddings, videos_embeddings)
        Z_repr, A_lg_strict, A_lg_soft = self.crossv(videos_embeddings, F_repr)
        out = self.mlp(torch.cat([F_repr, Z_repr], dim=-1))
        T = self.layernormT(self.pos_emb(self.T_embedding(x.permute(0, 2, 1))))
        T = self.layernormT1(self.attn1(T, out))
        pred = self.fc1(T).permute(0, 2, 1)
        pred_denorm = self.revin_layer(pred, 'denorm')
        if y is not None:
            loss_pred = self.criterion(pred_denorm, y.to(pred_denorm.device))
        else:
            loss_pred = torch.tensor(0.0, device=pred_denorm.device)
        def _flatten_tokens(x3): return x3.reshape(x3.size(0)*x3.size(1), x3.size(2))
        F_tok_n = F.normalize(_flatten_tokens(F_repr), dim=-1)
        Z_tok_n = F.normalize(_flatten_tokens(Z_repr), dim=-1)
        logits = torch.matmul(F_tok_n, Z_tok_n.t()) / self.mi_tau
        labels = torch.arange(logits.size(0), device=logits.device)
        L_mi_div = -F.cross_entropy(logits, labels)
        diff_hist = A_gl_strict - A_lg_strict.transpose(-1, -2)
        E_causal = (diff_hist * diff_hist).mean()
        L_hist = L_mi_div + self.lambda_causal * (-E_causal)
        eps = 1e-8
        P = A_gl_soft.clamp_min(eps); P /= P.sum(dim=-1, keepdim=True)
        Q = A_lg_soft.transpose(-1, -2).clamp_min(eps); Q /= Q.sum(dim=-1, keepdim=True)
        kl_pq = (P * (P.add(eps).log() - Q.add(eps).log())).sum(dim=-1).mean()
        kl_qp = (Q * (Q.add(eps).log() - P.add(eps).log())).sum(dim=-1).mean()
        L_kl_future = 0.5 * (kl_pq + kl_qp)
        L_repr_future = F.mse_loss(self.align_proj(F_repr), self.align_proj(Z_repr))
        L_future = self.alpha_align * L_kl_future + (1.0 - self.alpha_align) * L_repr_future
        L_cdia = L_hist + self.eta_future * L_future
        total_loss = self.w_cdia * L_cdia
        self.last_losses = {
            'pred': float(loss_pred.detach().item()),
            'hist_mi_div': float(L_mi_div.detach().item()),
            'hist_negE': float((-E_causal).detach().item()),
            'future_kl': float(L_kl_future.detach().item()),
            'future_repr': float(L_repr_future.detach().item()),
            'cdia': float(L_cdia.detach().item()),
            'total': float(total_loss.detach().item())
        }

        # ---- Excel 实时保存 ----
        self._log_step += 1
        self._append_losses_to_excel(step=self._log_step)

        return pred_denorm, total_loss * 0.1
